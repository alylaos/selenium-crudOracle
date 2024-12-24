from selenium import webdriver
from selenium.webdriver.support.ui import Select
import selenium.webdriver.support.ui as ui
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time 
from decimal import Decimal
from datetime import date, datetime, timedelta
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.remote.command import Command



browser=webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
#browser = webdriver.Chrome(r"C:\bitbucket\rpa_mysales\migracion mysales\chromedriver")

#opts = Options()
#opts.add_argument(
#    "user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/71.0.3578.80 Chrome/71.0.3578.80 Safari/537.36")
#opts.add_argument("--start-maximized")
#opts.add_argument("--ignore-certificate-errors")
#s = Service('chromedriver.exe')

book = load_workbook(r"CARGA BOT MINTRA-BAMBAS.xlsx")
sheet = book.active
timeout = 10
#browser = webdriver.Chrome(service=s, options=opts)

main_page = browser.current_window_handle
#book = load_workbook(r"data_mintra.xlsx")
#sheet = book.active

#inicio de sesion
browser.get('https://api-seguridad.sunat.gob.pe/v1/clientessol/b3639111-1546-4d06-b74f-de2c40629748/oauth2/login?originalUrl=https://apps.trabajo.gob.pe/si.segurovida/index.jsp&state=m1ntr4')
ruc = browser.find_element(By.ID,'txtRuc')
ruc.send_keys('20422293699')
usuario = browser.find_element(By.ID,'txtUsuario')
usuario.send_keys('IN2SCXLB')
clave = browser.find_element(By.ID,'txtContrasena')
clave.send_keys('Allied24')
botonlogin = browser.find_element(By.ID,'btnAceptar')
botonlogin.click()
#fin inicio sesion

time.sleep(10)

#nuevopaso

#ruc2 = browser.find_element(By.XPATH,'/html/body/form/table/tbody/tr[1]/td[2]/input').clear()
#ruc2 = browser.find_element(By.XPATH,'/html/body/form/table/tbody/tr[1]/td[2]/input')
#ruc2.send_keys('20422293699')
#time.sleep(3)

#aceptar_poliza = browser.find_element(By.XPATH,'/html/body/form/table/tbody/tr[3]/td/input')
#aceptar_poliza.click()
#time.sleep(3)

### click en actualizar, renovar poliza
boton_editar = browser.find_element(By.XPATH,'/html/body/center/div/div/div/map/area[2]')
boton_editar.click()
time.sleep(3)


tipo_poliza = browser.find_element(By.XPATH,'/html/body/center/table/tbody/tr/td/div/div[2]/form/fieldset/div[1]/div[2]/select')
tipo_poliza.click()
time.sleep(3)

##RIMAC
tipo_poliza = browser.find_element(By.XPATH,'/html/body/center/table/tbody/tr/td/div/div[2]/form/fieldset/div[1]/div[2]/select/option[4]')
tipo_poliza.click()
time.sleep(3)


numero_poliza = browser.find_element(By.XPATH,'/html/body/center/table/tbody/tr/td/div/div[2]/form/fieldset/div[1]/div[4]/input')
numero_poliza.send_keys('43151300170496')
time.sleep(3)

##ACEPTAR
aceptar_poliza = browser.find_element(By.XPATH,'/html/body/center/table/tbody/tr/td/div/div[2]/form/fieldset/div[2]/div/a')
aceptar_poliza.click()
time.sleep(3)

##radio button
elegir_poliza = browser.find_element(By.XPATH,'/html/body/center/table/tbody/tr/td/div/div[2]/form/div[1]/div/table/tbody/tr/td[1]/center/input')
elegir_poliza.click()
time.sleep(3)


##ALTAS Y BAJAS
altas_btn = browser.find_element(By.XPATH,'/html/body/center/table/tbody/tr/td/div/div[2]/form/div[2]/div/fieldset/a[1]')
altas_btn.click()
time.sleep(3)


#recorremos el excel
for i in range(1,sheet.max_row):
    #tipo_documento = sheet.cell(row=i+1, column=1).value
    documento = sheet.cell(row=i+1, column=1).value
    motivo = sheet.cell(row=i+1, column=2).value
    fecha = sheet.cell(row=i+1, column=3).value
    fecha_texto = fecha.strftime('%d%m%Y')


    numero_documento = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[5]/div[1]/input')
    numero_documento.send_keys(documento)

    lupa = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[5]/div[1]/label/a/img')
    lupa.click()
    time.sleep(5)

    baja = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[7]/div/table/tbody/tr/td/table/tbody/tr[1]/td[10]/a')
    baja.click()
    time.sleep(5)

    #nueva ventana de  baja
    for handle in browser.window_handles:
        if handle != main_page:
            nueva_ventana = handle
    
    browser.switch_to.window(nueva_ventana)
    ##ESCOGER motivo
    

    tipo_motivo = browser.find_element(By.XPATH,'/html/body/form/center/table/tbody/tr/td/table[2]/tbody/tr[3]/td/select')
    tipo_motivo.click()
    time.sleep(3)


    if motivo == '01': #renuncia
        tipo_motivo = browser.find_element(By.XPATH,'/html/body/form/center/table/tbody/tr/td/table[2]/tbody/tr[3]/td/select/option[1]')
    elif motivo == '09': #fallecimiento
        tipo_motivo = browser.find_element(By.XPATH,'/html/body/form/center/table/tbody/tr/td/table[2]/tbody/tr[3]/td/select/option[9]')
    
    tipo_motivo.click()
    time.sleep(5)
    

    #fecha
    fecha_input = browser.find_element(By.XPATH,'/html/body/form/center/table/tbody/tr/td/table[2]/tbody/tr[5]/td/input')
    fecha_input.send_keys(fecha_texto)
    time.sleep(5)

    #declaro
    seguro = browser.find_element(By.XPATH,'/html/body/form/center/table/tbody/tr/td/table[2]/tbody/tr[7]/td/input')
    seguro.click()
    time.sleep(5)

    #grabar
    grabar = browser.find_element(By.XPATH,'/html/body/form/center/table/tbody/tr/td/table[2]/tbody/tr[8]/td/a[1]')
    grabar.click()
    print("se registró")
    time.sleep(10)
    alert = browser.switch_to.alert
    alert.accept()
    print("se aceptò")
    time.sleep(10)

    print(documento)

    for handle in browser.window_handles:
        if handle != main_page:
            nueva_ventana = handle
    
    browser.switch_to.window(nueva_ventana)
    print("se cambio de ventana")

    numero_documento = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[5]/div[1]/input')
    numero_documento.clear()
    time.sleep(5)
    print("se limpió")
    
    #aceptar la baja

    #browser.switchTo().alert().accept();
    #time.sleep(5)
    
      
   
    
    

    


    




    

    


    






















# fila_serv = 1
# cliente_cotizacion = ''
# parent = 1
# child = 1
# indice = 1
# #recorremos el excel
# for i in range(1,sheet.max_row):
#     #obtenemos las celdas del excel
#     nomcliente = sheet.cell(row=i+1, column=1).value
#     accesorio = sheet.cell(row=i+1, column=2).value
#     notas_accesorio = sheet.cell(row=i+1, column=3).value
#     importe_acc = sheet.cell(row=i+1, column=4).value

#     #validamos si la siguiente linea del excel pertenece al mismo cliente
#     if cliente_cotizacion != nomcliente:
#         #inicio crear nueva cotizacion
#         browser.get('https://mysales-uat.pe.g4s.com/ventas/cotizacion/buscar')
#         time.sleep(3)
#         nueva_cotizacion = browser.find_element(By.ID,'btnNuevaCotSSFF')
#         nueva_cotizacion.click()
#         time.sleep(3)
#         moneda_cotizacion = browser.find_element(By.ID,'aceptarBtn')
#         moneda_cotizacion.click()
#         time.sleep(3)
#         cliente_texto = browser.find_element(By.ID,'id_razon_social')
#         cliente_texto.clear()
#         cliente_texto.send_keys(nomcliente)
#         time.sleep(3)
#         cliente_texto.send_keys(Keys.ARROW_DOWN)
#         cliente_texto.send_keys(Keys.ENTER)
#         tiempo_contrato = browser.find_element(By.ID,'id_tiempo_servicio')
#         fecha_estimada = browser.find_element(By.ID,'id_fecha_inicio_estimada')
#         tiempo_contrato.clear()
#         tiempo_contrato.send_keys('12')
#         fecha_estimada.click()
#         fecha_estimada2 = browser.find_element(By.CLASS_NAME,'day')
#         fecha_estimada2.click()
#         tiempo_contrato.click()
#         indice = 1
#         #fin crear nueva cotizacion
#     cliente_cotizacion = nomcliente
#     btn_accesorios = browser.find_element(By.ID,'accesorios')
#     btn_accesorios.click()
#     time.sleep(3)
#     btn_accesorios_add = browser.find_element(By.ID,'accesorios_agregar')
#     btn_accesorios_add.click()
#     time.sleep(2)
#     input_accesorios = browser.find_element(By.ID,'nombre_carticulo_acc_%s' % indice)
#     input_accesorios.send_keys(accesorio)
#     time.sleep(30)

#     input_accesorios.send_keys(Keys.ARROW_DOWN)
#     input_accesorios.send_keys(Keys.ENTER)
#     input_nota_accesorios = browser.find_element(By.ID,'descripcion_carticulo_acc_%s' % indice)
#     input_nota_accesorios.send_keys(notas_accesorio)

#     input_vigencia_accesorios = browser.find_element(By.ID,'vigencia_carticulo_acc_%s' % indice)
#     val_vigencia = input_vigencia_accesorios.get_attribute("data-vigencia")

#     input_nuevoprecio_accesorios = browser.find_element(By.ID,'nuevo_precio_carticulo_acc_%s' % indice)
#     ant_precio = Decimal(input_nuevoprecio_accesorios.get_attribute("value"))
#     nuevo_precio = int(val_vigencia)*Decimal(importe_acc)
#     if nuevo_precio >= ant_precio:
#         input_nuevoprecio_accesorios.clear()
#         time.sleep(3)
#         input_nuevoprecio_accesorios.send_keys(str(nuevo_precio))
#         time.sleep(3)
#         btn_save_accesorios = browser.find_element(By.ID,'accesorios_guardar')
#         btn_save_accesorios.click()
#         time.sleep(3)
#         btn_quit_accesorios = browser.find_element(By.ID,'accesorios_salir')
#         btn_quit_accesorios.click()
#         time.sleep(3)
#         btn_save_cotizacion = browser.find_element(By.ID,'cotizacion_guardar')
#         btn_save_cotizacion.click()

#     time.sleep(3)
#     if i+1 == sheet.max_row:
#         btn_save_accesorios = browser.find_element(By.ID,'accesorios_guardar')
#         btn_save_accesorios.click()
#         time.sleep(3)
#         btn_quit_accesorios = browser.find_element(By.ID,'accesorios_salir')
#         btn_quit_accesorios.click()
#         time.sleep(3)
#         btn_save_cotizacion = browser.find_element(By.ID,'cotizacion_guardar')
#         btn_save_cotizacion.click()
#         time.sleep(3)

#     indice+=1