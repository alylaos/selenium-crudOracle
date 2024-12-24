from collections import namedtuple
from pickle import NONE
from sqlite3 import Cursor
from setting import settings as sett
from time import sleep
import os
import time
from datetime import datetime   
from model import ALTAS_CESES
import openpyxl
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import pickle
from openpyxl import load_workbook
from datetime import date 
import smtplib
from email.message import EmailMessage
import imghdr
from datetime import date 


EMAIL_ADDRESS = "bot1@pe.g4s.com"
EMAIL_PASSWORD = "cyra xkqh xplf gbjm"
email = 0
email_date = date.today()

SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly','https://www.googleapis.com/auth/drive','https://www.googleapis.com/auth/spreadsheets']
SAMPLE_SPREADSHEET_ID = '1AM_jNx50ykF4zLufE-zE-e7reTxiDJyMQXWHENsT274' 
SAMPLE_RANGE_NAME = 'CESES!A2:R' ## CAMBIAR NOMBRE Y NUMERO
SAMPLE_RANGE_NAME_2 = 'ALTAS!A2:T'
SAMPLE_RANGE_NAME_3 = 'FECHAS!A2:B'

ruta = 'E:\\Pentaho\\Jobs\\Reporte_Altas_Ceses\\Documents\\Altas\\lista_altas.xlsx'
ruta2 = 'E:\\Pentaho\\Jobs\\Reporte_Altas_Ceses\\Documents\\Ceses\\lista_ceses.xlsx'
ruta3 ='E:\\Pentaho\\Jobs\\Reporte_Altas_Ceses\\credenciales\\token.pickle'
ruta4 = 'E:\\Pentaho\\Jobs\\Reporte_Altas_Ceses\\\\credenciales\\credentials.json'
ruta5 = 'E:\\Pentaho\\Jobs\\Reporte_Altas_Ceses\\Documents\\Fechas\\fechas.xlsx'
ruta6 = 'E:\\Pentaho\\Jobs\\Reporte_Altas_Ceses\\Documents\\Altas'
ruta7 = 'E:\\Pentaho\\Jobs\\Reporte_Altas_Ceses\\log_ALTAS_CESES.txt'
ruta8 = 'E:\\Pentaho\\Jobs\\Reporte_Altas_Ceses\\Documents\\Ceses'

fecha_inicio = '26/12/2021'
fecha_fin = date.today()
fecha_fin = str(fecha_fin)
fecha_fin =  fecha_fin[8:10]+"/"+fecha_fin[5:7]+"/"+fecha_fin[0:4]
#fecha_fin = '22/08/2022'
##fecha_fin = '08/08/2022'

def listar_ceses():
    carpeta = os.listdir(ruta8)
    print(f"files: {carpeta}")
    if len(carpeta) > 0:
        os.remove(ruta2)
        print("x")
    datos = ALTAS_CESES.listar_datos_ceses(fecha_inicio,fecha_fin)
    print("COUNT: "+str(len(datos)))
    wb = openpyxl.Workbook()
    sheet = wb.active
    band = 1 
    fila = 1
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11, i12, i13, i14, i15, i16, i17, i18  in datos:
        
        if band==1:
            c2 = sheet.cell(row= fila , column = 1).value = "Matricula"
            c2 = sheet.cell(row= fila , column = 2).value = "Documento"
            c2 = sheet.cell(row= fila , column = 3).value = "Apellidos y Nombres"
            c2 = sheet.cell(row= fila , column = 4).value = "Tipo Trabajador"
            c2 = sheet.cell(row= fila , column = 5).value = "Fecha ingreso"
            c2 = sheet.cell(row= fila , column = 6).value = "Situacion trabajador"
            c2 = sheet.cell(row= fila , column = 7).value = "Fecha de Retiro"
            c2 = sheet.cell(row= fila , column = 8).value = "Email Trabajo"
            c2 = sheet.cell(row= fila , column = 9).value = "Email Personal"
            c2 = sheet.cell(row= fila , column = 10).value = "Telefono movil"
            c2 = sheet.cell(row= fila , column = 11).value = "Unidad Funcional"
            c2 = sheet.cell(row= fila , column = 12).value = "Ubicación Fisica"
            c2 = sheet.cell(row= fila , column = 13).value = "Descripcion Puesto"
            c2 = sheet.cell(row= fila , column = 14).value = "Id Area"
            c2 = sheet.cell(row= fila , column = 15).value = "Codigo Cliente G4S"
            c2 = sheet.cell(row= fila , column = 16).value = "Nombre Cliente G4S"
            c2 = sheet.cell(row= fila , column = 17).value = "Unidad negocio"
            c2 = sheet.cell(row= fila , column = 18).value = "Sucursal"
            band=0
            fila = fila +1
              
        i1 = str(i1)
        c2 = sheet.cell(row= fila , column = 1).value = i1
        c2 = sheet.cell(row= fila , column = 2).value = i2
        c2 = sheet.cell(row= fila , column = 3).value = i3
        c2 = sheet.cell(row= fila , column = 4).value = i4
        i5 = datetime.date(i5)
        i5 = str(i5)
        i5 = i5[8:10]+"/"+i5[5:7]+"/"+i5[0:4]
        c2 = sheet.cell(row= fila , column = 5).value = i5
        c2 = sheet.cell(row= fila , column = 6).value = i6
        i7 = datetime.date(i7)
        i7 = str(i7)
        i7 = i7[8:10]+"/"+i7[5:7]+"/"+i7[0:4]
        c2 = sheet.cell(row= fila , column = 7).value = i7
        c2 = sheet.cell(row= fila , column = 8).value = i8
        c2 = sheet.cell(row= fila , column = 9).value = i9
        c2 = sheet.cell(row= fila , column = 10).value = i10
        i11 = str(i11)
        c2 = sheet.cell(row= fila , column = 11).value = i11
        c2 = sheet.cell(row= fila , column = 12).value = i12
        c2 = sheet.cell(row= fila , column = 13).value = i13
        c2 = sheet.cell(row= fila , column = 14).value = i14
        c2 = sheet.cell(row= fila , column = 15).value = i15
        c2 = sheet.cell(row= fila , column = 16).value = i16
        c2 = sheet.cell(row= fila , column = 17).value = i17
        c2 = sheet.cell(row= fila , column = 18).value = i18
        fila = fila + 1

    wb.save(ruta2) 
    print("REPORTE DE CESES LISTO.")


def listar_altas():
    datos2 = ALTAS_CESES.listar_datos_altas(fecha_inicio,fecha_fin)
    print("Count: "+str(len(datos2)))
    wb = openpyxl.Workbook()
    sheet = wb.active
    band = 1 
    fila = 1
    
    for i1, i2, i3, i4, i5, i6, i7, i8, i9, i10, i11, i12, i13, i14, i15, i16, i17, i18, i19  in datos2:
        if band==1:
            c2 = sheet.cell(row= fila , column = 1).value = "Matricula"
            c2 = sheet.cell(row= fila , column = 2).value = "NRO RQ"
            c2 = sheet.cell(row= fila , column = 3).value = "Documento"
            c2 = sheet.cell(row= fila , column = 4).value = "Apellidos y Nombres"
            c2 = sheet.cell(row= fila , column = 5).value = "Tipo Trabajador"
            c2 = sheet.cell(row= fila , column = 6).value = "Fecha ingreso"
            c2 = sheet.cell(row= fila , column = 7).value = "Situacion trabajador"
            c2 = sheet.cell(row= fila , column = 8).value = "Fecha de Retiro"
            c2 = sheet.cell(row= fila , column = 9).value = "Email Trabajo"
            c2 = sheet.cell(row= fila , column = 10).value = "Email Personal"
            c2 = sheet.cell(row= fila , column = 11).value = "Telefono movil"
            c2 = sheet.cell(row= fila , column = 12).value = "Unidad Funcional"
            c2 = sheet.cell(row= fila , column = 13).value = "Ubicación Fisica"
            c2 = sheet.cell(row= fila , column = 14).value = "Descripcion Puesto"
            c2 = sheet.cell(row= fila , column = 15).value = "Id Area"
            c2 = sheet.cell(row= fila , column = 16).value = "Codigo Cliente G4S"
            c2 = sheet.cell(row= fila , column = 17).value = "Nombre Cliente G4S"
            c2 = sheet.cell(row= fila , column = 18).value = "Unidad negocio"
            c2 = sheet.cell(row= fila , column = 19).value = "Analista Seleccion"
            band=0
            fila = fila +1
              
        i1 = str(i1)
        c2 = sheet.cell(row= fila , column = 1).value = i1
        c2 = sheet.cell(row= fila , column = 2).value = i2
        c2 = sheet.cell(row= fila , column = 3).value = i3
        c2 = sheet.cell(row= fila , column = 4).value = i4
        c2 = sheet.cell(row= fila , column = 5).value = i5
        
        if i6 is None:
            c2 = sheet.cell(row= fila , column = 6).value = i6
        else:
            i6 = datetime.date(i6)
            i6 = str(i6)
            i6 = i6[8:10]+"/"+i6[5:7]+"/"+i6[0:4]
            c2 = sheet.cell(row= fila , column = 6).value = i6

        c2 = sheet.cell(row= fila , column = 7).value = i7        
        
        if i8 is None:
            c2 = sheet.cell(row= fila , column = 8).value = i8
        else:
            i8 = datetime.date(i8)
            i8 = str(i8)
            i8 = i8[8:10]+"/"+i8[5:7]+"/"+i8[0:4]
            c2 = sheet.cell(row= fila , column = 8).value = i8
       
        c2 = sheet.cell(row= fila , column = 9).value = i9
        c2 = sheet.cell(row= fila , column = 10).value = i10
        c2 = sheet.cell(row= fila , column = 11).value = i11
        c2 = sheet.cell(row= fila , column = 12).value = i12
        c2 = sheet.cell(row= fila , column = 13).value = i13
        c2 = sheet.cell(row= fila , column = 14).value = i14
        c2 = sheet.cell(row= fila , column = 15).value = i15
        c2 = sheet.cell(row= fila , column = 16).value = i16
        c2 = sheet.cell(row= fila , column = 17).value = i17
        c2 = sheet.cell(row= fila , column = 18).value = i18
        c2 = sheet.cell(row= fila , column = 19).value = i19
        fila = fila + 1

    wb.save(ruta) 
    print("REPORTE DE ALTAS LISTO.")


def cargar_ceses():
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists(ruta3):
        with open(ruta3, 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                ruta4, SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open(ruta3, 'wb') as token:
            pickle.dump(creds, token)

    service = build('sheets', 'v4', credentials=creds)

    # Call the Sheets API
    sheet = service.spreadsheets()

    ###### LIMPIAR RANGO #######
    clear_values_request_body = {}
    request = sheet.values().clear(spreadsheetId=SAMPLE_SPREADSHEET_ID, range=SAMPLE_RANGE_NAME, body=clear_values_request_body)
    response = request.execute()

    # ruta de nuestro archivo
    filesheet = ruta2
    # creamos ell obejeto load_workbook
    wb = load_workbook(filesheet)
    # seleccionamos el archivo
    ws = wb.active
    
    rng=ws['A2':'R' + str(ws.max_row)]
    final_list=[]
    i=1
    for i1,i2,i3,i4,i5,i6,i7,i8,i9,i10,i11,i12,i13,i14,i15,i16,i17, i18 in rng:  
        final_list.append([i1.value,i2.value,"'%s" % i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value,i12.value,i13.value,i14.value,i15.value,i16.value,i17.value,i18.value])

    data = final_list
    #### ESCRIBIR EN EL DRIVE #######        
    value_input_option = 'USER_ENTERED'

    body = {
        'values': data
    }
    result = sheet.values().update(
        spreadsheetId=SAMPLE_SPREADSHEET_ID, range=SAMPLE_RANGE_NAME,
        valueInputOption=value_input_option, body=body).execute()
    print('CARGA DE CESES - {0} cells updated.'.format(result.get('updatedCells')))    


def cargar_altas():
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists(ruta3):
        with open(ruta3, 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                ruta4, SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open(ruta3, 'wb') as token:
            pickle.dump(creds, token)

    service = build('sheets', 'v4', credentials=creds)

    # Call the Sheets API
    sheet = service.spreadsheets()

    ###### LIMPIAR RANGO #######
    clear_values_request_body = {}
    request = sheet.values().clear(spreadsheetId=SAMPLE_SPREADSHEET_ID, range=SAMPLE_RANGE_NAME_2, body=clear_values_request_body)
    response = request.execute()

    # ruta de nuestro archivo
    filesheet = ruta
    # creamos ell obejeto load_workbook
    wb = load_workbook(filesheet)
    # seleccionamos el archivo
    ws = wb.active
    rng=ws['A1':'T' + str(ws.max_row)]
    final_list=[]
    i=1
    for i1,i2,i3,i4,i5,i6,i7,i8,i9,i10,i11,i12,i13,i14,i15,i16,i17, i18, i19, i20 in rng:  
        final_list.append([i1.value,i2.value,i3.value,i4.value,i5.value,i6.value,i7.value,i8.value,i9.value,i10.value,i11.value,i12.value,i13.value,i14.value,i15.value,i16.value,i17.value,i18.value,i19.value,i20.value])

    data = final_list
    #### ESCRIBIR EN EL DRIVE #######        
    value_input_option = 'USER_ENTERED'

    body = {
        'values': data
    }
    result = sheet.values().update(
        spreadsheetId=SAMPLE_SPREADSHEET_ID, range=SAMPLE_RANGE_NAME_2,
        valueInputOption=value_input_option, body=body).execute()
    print('CARGA DE ALTAS - {0} cells updated.'.format(result.get('updatedCells')))    


def listar_fecha():
    wb = openpyxl.Workbook()
    sheet = wb.active
    fila = 1
   
    c2 = sheet.cell(row= fila , column = 1).value = "Fecha Inicio"
    c2 = sheet.cell(row= fila , column = 2).value = "Fecha Final"
    fila = fila +1
    c2 = sheet.cell(row= fila , column = 1).value = fecha_inicio
    c2 = sheet.cell(row= fila , column = 2).value = fecha_fin
    wb.save(ruta5)     


def cargar_fecha():
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists(ruta3):
        with open(ruta3, 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                ruta4, SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open(ruta3, 'wb') as token:
            pickle.dump(creds, token)

    service = build('sheets', 'v4', credentials=creds)

    # Call the Sheets API
    sheet = service.spreadsheets()

    ###### LIMPIAR RANGO #######
    clear_values_request_body = {}
    request = sheet.values().clear(spreadsheetId=SAMPLE_SPREADSHEET_ID, range=SAMPLE_RANGE_NAME_3, body=clear_values_request_body)
    response = request.execute()

    # ruta de nuestro archivo
    filesheet = ruta5
    # creamos ell obejeto load_workbook
    wb = load_workbook(filesheet)
    # seleccionamos el archivo
    ws = wb.active
    rng=ws['A2':'B' + str(ws.max_row)]
    final_list=[]
    i=1
    for i1,i2 in rng:  
        final_list.append([i1.value,i2.value])

    data = final_list
    #### ESCRIBIR EN EL DRIVE #######        
    value_input_option = 'USER_ENTERED'

    body = {
        'values': data
    }
    result = sheet.values().update(
        spreadsheetId=SAMPLE_SPREADSHEET_ID, range=SAMPLE_RANGE_NAME_3,
        valueInputOption=value_input_option, body=body).execute()


def send_email():
    if email == 1:  
        msg = EmailMessage()
        msg['Subject'] = 'REPORTE DE ALTAS Y CESES - TAREA COMPLETADA DE MANERA EXITOSA'
        msg['From'] = EMAIL_ADDRESS
        msg['To'] = "adrian.nunez@pe.g4s.com"

        msg.set_content('TAREA REALIZADA - FECHA ('+str(email_date)+')')

        files = [ruta7]

        for file in files:
                with open(file,'rb') as f:
                    file_data = f.read()
                    file_type = imghdr.what(f.name)
                    file_name = f.name
                    x = file_name.split("\\")
                    file_name = x[4]
            
                msg.add_attachment(file_data, maintype='text', subtype='plain', filename=file_name)

        with smtplib.SMTP_SSL('smtp.gmail.com',465) as smtp:
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            smtp.send_message(msg)
        
        print("EMAIL ENVIADO")
    else:
        msg = EmailMessage()
        msg['Subject'] = 'REPORTE DE ALTAS Y CESES - TAREA NO COMPLETADA'
        msg['From'] = EMAIL_ADDRESS
        msg['To'] = "adrian.nunez@pe.g4s.com"
        msg.set_content('REVISAR - FECHA ('+str(email_date)+')')

        files = [ruta7]

        for file in files:
                with open(file,'rb') as f:
                    file_data = f.read()
                    file_type = imghdr.what(f.name)
                    file_name = f.name
                    x = file_name.split("\\")
                    file_name = x[4]
            
                msg.add_attachment(file_data, maintype='text', subtype='plain', filename=file_name)

        with smtplib.SMTP_SSL('smtp.gmail.com',465) as smtp:
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            smtp.send_message(msg)
        
        print("EMAIL ENVIADO")


if __name__ == '__main__':
    #listar_altas()
    try:
        listar_ceses()
        cargar_altas()
        cargar_ceses()
        listar_fecha()
        cargar_fecha()
        email = 1
        send_email()
    except:
        send_email()



