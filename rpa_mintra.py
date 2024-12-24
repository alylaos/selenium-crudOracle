from selenium import webdriver
from selenium.webdriver.support.ui import Select
import selenium.webdriver.support.ui as ui
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time 
from decimal import Decimal
from datetime import date, datetime, timedelta
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoAlertPresentException
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.common.exceptions import NoSuchElementException
 

browser=webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
#browser = webdriver.Chrome(r"C:\bitbucket\rpa_mysales\migracion mysales\chromedriver")

#opts = Options()
#opts.add_argument(
#    "user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/71.0.3578.80 Chrome/71.0.3578.80 Safari/537.36")
#opts.add_argument("--start-maximized")
#opts.add_argument("--ignore-certificate-errors")
#s = Service('chromedriver.exe')

book = load_workbook(r"CARGA BOT MINTRA-GENERAL.xlsx")
sheet = book.active
timeout = 10

#browser = webdriver.Chrome(service=s)

#book = load_workbook(r"data_mintra.xlsx")
#sheet = book.active

#inicio de sesion
browser.get('https://api-seguridad.sunat.gob.pe/v1/clientessol/b3639111-1546-4d06-b74f-de2c40629748/oauth2/login?originalUrl=https://apps.trabajo.gob.pe/si.segurovida/index.jsp&state=m1ntr4')
ruc = browser.find_element(By.ID,'txtRuc')
ruc.send_keys('20422293699')
usuario = browser.find_element(By.ID,'txtUsuario')
usuario.send_keys('IN2SCXLB')
clave = browser.find_element(By.ID,'txtContrasena')
clave.send_keys('Allied24')
botonlogin = browser.find_element(By.ID,'btnAceptar')
botonlogin.click()
#fin inicio sesion

time.sleep(20)


### click en actualizar, renovar poliza
boton_editar = browser.find_element(By.XPATH,'/html/body/center/div/div/div/map/area[2]')
boton_editar.click()
time.sleep(3)


tipo_poliza = browser.find_element(By.XPATH,'/html/body/center/table/tbody/tr/td/div/div[2]/form/fieldset/div[1]/div[2]/select')
tipo_poliza.click()
time.sleep(3)

##RIMAC
tipo_poliza = browser.find_element(By.XPATH,'/html/body/center/table/tbody/tr/td/div/div[2]/form/fieldset/div[1]/div[2]/select/option[4]')
tipo_poliza.click()
time.sleep(3)


numero_poliza = browser.find_element(By.XPATH,'/html/body/center/table/tbody/tr/td/div/div[2]/form/fieldset/div[1]/div[4]/input')
numero_poliza.send_keys('43151300170498')
time.sleep(3)

##ACEPTAR
aceptar_poliza = browser.find_element(By.XPATH,'/html/body/center/table/tbody/tr/td/div/div[2]/form/fieldset/div[2]/div/a')
aceptar_poliza.click()
time.sleep(3)

##radio button
elegir_poliza = browser.find_element(By.XPATH,'/html/body/center/table/tbody/tr/td/div/div[2]/form/div[1]/div/table/tbody/tr/td[1]/center/input')
elegir_poliza.click()
time.sleep(3)


##ALTAS Y BAJAS
altas_btn = browser.find_element(By.XPATH,'/html/body/center/table/tbody/tr/td/div/div[2]/form/div[2]/div/fieldset/a[1]')
altas_btn.click()
time.sleep(3)


#recorremos el excel
for i in range(1,sheet.max_row):
	tipo_documento = sheet.cell(row=i+1, column=1).value
	documento = sheet.cell(row=i+1, column=2).value
	fecha = sheet.cell(row=i+1, column=3).value
	fecha_texto = fecha.strftime('%d%m%Y')
	monto = sheet.cell(row=i+1, column=4).value
	try:
		numero_documento = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[1]/div[1]/div[2]/input')
		numero_documento.clear()
		time.sleep(5)

		fecha_input = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[2]/div[1]/div[3]/input')
		fecha_input.clear()
		time.sleep(5)

		monto_input = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[2]/div[2]/div[3]/input')
		monto_input.clear()
		time.sleep(5)
	
		tipo_doc = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[1]/div[1]/div[1]/select')
		tipo_doc.click()
		time.sleep(3)

		##ESCOGER DOCUMENTO
		if tipo_documento == '03':
			tipo_doc = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[1]/div[1]/div[1]/select/option[2]')
		elif tipo_documento == '06':
			tipo_doc = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[1]/div[1]/div[1]/select/option[3]')
		elif tipo_documento == '08':
			tipo_doc = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[1]/div[1]/div[1]/select/option[4]')
		elif tipo_documento == '17':
			tipo_doc = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[1]/div[1]/div[1]/select/option[5]')
		elif tipo_documento == '18':
			tipo_doc = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[1]/div[1]/div[1]/select/option[6]')
		elif tipo_documento == '26':
			tipo_doc = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[1]/div[1]/div[1]/select/option[7]')
		else:
			tipo_doc = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[1]/div[1]/div[1]/select/option[2]')

		tipo_doc.click()

		numero_documento = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[1]/div[1]/div[2]/input')
		numero_documento.send_keys(documento)
		numero_documento.send_keys(Keys.ENTER)

		time.sleep(5)

		##ESCOGER SEGURO
		seguro = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[2]/div[1]/div[1]/input[2]')
		seguro.click()
		time.sleep(5)	

		##ESCOGER SEGURO
		seguro = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[2]/div[1]/div[2]/input[2]')
		seguro.click()
		time.sleep(5)

		fecha_input = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[2]/div[1]/div[3]/input')
		fecha_input.send_keys(fecha_texto)
		time.sleep(5)

		monto_input = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[2]/div[2]/div[3]/input')
		monto_input.send_keys(monto)
		time.sleep(5)

		grabar = browser.find_element(By.XPATH,'/html/body/div/form/div/div/div[3]/fieldset[2]/fieldset[2]/div[3]/div/a')
		grabar.click()
		print(documento)

		time.sleep(5)

	except UnexpectedAlertPresentException:
		WebDriverWait(browser, timeout).until(tipo_doc)
		alerta = browser.switch_to.alert 
		alerta.accept()
		time.sleep(10)
	
	except NoAlertPresentException:
    # Si no se encuentra ninguna alerta, continúa con el flujo normal del programa
		pass
		time.sleep(10)
		
	except	NoSuchElementException as e: 
		print(f"Elemento no encontrado: {e}")
		
	






	

	

	


	




	

	


	






















# fila_serv = 1
# cliente_cotizacion = ''
# parent = 1
# child = 1
# indice = 1
# #recorremos el excel
# for i in range(1,sheet.max_row):
#     #obtenemos las celdas del excel
#     nomcliente = sheet.cell(row=i+1, column=1).value
#     accesorio = sheet.cell(row=i+1, column=2).value
#     notas_accesorio = sheet.cell(row=i+1, column=3).value
#     importe_acc = sheet.cell(row=i+1, column=4).value

#     #validamos si la siguiente linea del excel pertenece al mismo cliente
#     if cliente_cotizacion != nomcliente:
#         #inicio crear nueva cotizacion
#         browser.get('https://mysales-uat.pe.g4s.com/ventas/cotizacion/buscar')
#         time.sleep(3)
#         nueva_cotizacion = browser.find_element(By.ID,'btnNuevaCotSSFF')
#         nueva_cotizacion.click()
#         time.sleep(3)
#         moneda_cotizacion = browser.find_element(By.ID,'aceptarBtn')
#         moneda_cotizacion.click()
#         time.sleep(3)
#         cliente_texto = browser.find_element(By.ID,'id_razon_social')
#         cliente_texto.clear()
#         cliente_texto.send_keys(nomcliente)
#         time.sleep(3)
#         cliente_texto.send_keys(Keys.ARROW_DOWN)
#         cliente_texto.send_keys(Keys.ENTER)
#         tiempo_contrato = browser.find_element(By.ID,'id_tiempo_servicio')
#         fecha_estimada = browser.find_element(By.ID,'id_fecha_inicio_estimada')
#         tiempo_contrato.clear()
#         tiempo_contrato.send_keys('12')
#         fecha_estimada.click()
#         fecha_estimada2 = browser.find_element(By.CLASS_NAME,'day')
#         fecha_estimada2.click()
#         tiempo_contrato.click()
#         indice = 1
#         #fin crear nueva cotizacion
#     cliente_cotizacion = nomcliente
#     btn_accesorios = browser.find_element(By.ID,'accesorios')
#     btn_accesorios.click()
#     time.sleep(3)
#     btn_accesorios_add = browser.find_element(By.ID,'accesorios_agregar')
#     btn_accesorios_add.click()
#     time.sleep(2)
#     input_accesorios = browser.find_element(By.ID,'nombre_carticulo_acc_%s' % indice)
#     input_accesorios.send_keys(accesorio)
#     time.sleep(30)

#     input_accesorios.send_keys(Keys.ARROW_DOWN)
#     input_accesorios.send_keys(Keys.ENTER)
#     input_nota_accesorios = browser.find_element(By.ID,'descripcion_carticulo_acc_%s' % indice)
#     input_nota_accesorios.send_keys(notas_accesorio)

#     input_vigencia_accesorios = browser.find_element(By.ID,'vigencia_carticulo_acc_%s' % indice)
#     val_vigencia = input_vigencia_accesorios.get_attribute("data-vigencia")

#     input_nuevoprecio_accesorios = browser.find_element(By.ID,'nuevo_precio_carticulo_acc_%s' % indice)
#     ant_precio = Decimal(input_nuevoprecio_accesorios.get_attribute("value"))
#     nuevo_precio = int(val_vigencia)*Decimal(importe_acc)
#     if nuevo_precio >= ant_precio:
#         input_nuevoprecio_accesorios.clear()
#         time.sleep(3)
#         input_nuevoprecio_accesorios.send_keys(str(nuevo_precio))
#         time.sleep(3)
#         btn_save_accesorios = browser.find_element(By.ID,'accesorios_guardar')
#         btn_save_accesorios.click()
#         time.sleep(3)
#         btn_quit_accesorios = browser.find_element(By.ID,'accesorios_salir')
#         btn_quit_accesorios.click()
#         time.sleep(3)
#         btn_save_cotizacion = browser.find_element(By.ID,'cotizacion_guardar')
#         btn_save_cotizacion.click()

#     time.sleep(3)
#     if i+1 == sheet.max_row:
#         btn_save_accesorios = browser.find_element(By.ID,'accesorios_guardar')
#         btn_save_accesorios.click()
#         time.sleep(3)
#         btn_quit_accesorios = browser.find_element(By.ID,'accesorios_salir')
#         btn_quit_accesorios.click()
#         time.sleep(3)
#         btn_save_cotizacion = browser.find_element(By.ID,'cotizacion_guardar')
#         btn_save_cotizacion.click()
#         time.sleep(3)

#     indice+=1